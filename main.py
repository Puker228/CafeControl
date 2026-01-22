import re
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, ttk
from tkinter.filedialog import asksaveasfilename
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import (
    ForeignKey,
    Numeric,
    create_engine,
    event,
    text,
)
from sqlalchemy.orm import (
    DeclarativeBase,
    Mapped,
    Session,
    mapped_column,
    relationship,
    sessionmaker,
)

# ===================== DATABASE =====================


class Base(DeclarativeBase):
    id: Mapped[int] = mapped_column(primary_key=True)


class Customer(Base):
    __tablename__ = "customers"

    name: Mapped[str]
    phone: Mapped[str]
    email: Mapped[str] = mapped_column(unique=True)
    loyalty_level: Mapped[str] = mapped_column(default="Bronze")
    discount_percent: Mapped[float] = mapped_column(default=0.0)

    orders: Mapped[list["Order"]] = relationship(
        back_populates="customer", cascade="all, delete"
    )


class Employee(Base):
    __tablename__ = "employees"
    fio: Mapped[str]
    role: Mapped[str]  # бармен, официант, администратор и тд
    phone: Mapped[str]
    hire_date: Mapped[datetime] = mapped_column(default=datetime.now)
    salary: Mapped[float] = mapped_column(Numeric(10, 2))

    orders: Mapped[list["Order"]] = relationship(back_populates="employee")


class Supplier(Base):
    __tablename__ = "suppliers"
    name: Mapped[str]
    phone: Mapped[str]
    email: Mapped[str]
    address: Mapped[str]

    ingredients: Mapped[list["Ingredient"]] = relationship(back_populates="supplier")


class Ingredient(Base):
    __tablename__ = "ingredients"
    name: Mapped[str]
    unit: Mapped[str]  # л, кг, шт
    stock_quantity: Mapped[float]
    min_stock_level: Mapped[float]
    purchase_price: Mapped[float] = mapped_column(Numeric(10, 2))

    supplier_id: Mapped[int] = mapped_column(ForeignKey("suppliers.id"))
    supplier: Mapped["Supplier"] = relationship(back_populates="ingredients")

    recipes: Mapped[list["Recipe"]] = relationship(back_populates="ingredient")


class MenuItem(Base):
    __tablename__ = "menu_items"
    name: Mapped[str] = mapped_column(unique=True)
    type: Mapped[str]  # еда / напиток / алкоголь ...
    selling_price: Mapped[float] = mapped_column(Numeric(10, 2))
    volume_or_weight: Mapped[str]

    compositions: Mapped[list["OrderComposition"]] = relationship(
        back_populates="menu_item"
    )
    recipes: Mapped[list["Recipe"]] = relationship(back_populates="menu_item")


class Recipe(Base):
    __tablename__ = "recipes"
    menu_item_id: Mapped[int] = mapped_column(ForeignKey("menu_items.id"))
    ingredient_id: Mapped[int] = mapped_column(ForeignKey("ingredients.id"))
    quantity_required: Mapped[float]
    unit: Mapped[str]

    menu_item: Mapped["MenuItem"] = relationship(back_populates="recipes")
    ingredient: Mapped["Ingredient"] = relationship(back_populates="recipes")


class Order(Base):
    __tablename__ = "orders"

    order_date: Mapped[datetime] = mapped_column(default=datetime.now)
    order_type: Mapped[str]
    total_amount: Mapped[float] = mapped_column(Numeric(10, 2), default=0.0)
    payment_method: Mapped[str]
    status: Mapped[str]

    customer_id: Mapped[Optional[int]] = mapped_column(
        ForeignKey("customers.id"), nullable=True
    )
    customer: Mapped[Optional["Customer"]] = relationship(back_populates="orders")

    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    employee: Mapped["Employee"] = relationship(back_populates="orders")

    compositions: Mapped[list["OrderComposition"]] = relationship(
        back_populates="order", cascade="all, delete"
    )

    def total(self):
        return round(sum(i.total_price() for i in self.compositions), 2)


class OrderComposition(Base):
    __tablename__ = "order_compositions"

    order_id: Mapped[int] = mapped_column(ForeignKey("orders.id"))
    order: Mapped["Order"] = relationship(back_populates="compositions")

    menu_item_id: Mapped[int] = mapped_column(ForeignKey("menu_items.id"))
    menu_item: Mapped["MenuItem"] = relationship(back_populates="compositions")

    quantity: Mapped[int] = mapped_column(default=1)
    price_at_sale: Mapped[float] = mapped_column(Numeric(10, 2))

    def total_price(self):
        return round(float(self.price_at_sale) * self.quantity, 2)


class TriggerLog(Base):
    __tablename__ = "trigger_logs"

    trigger_name: Mapped[str]
    action: Mapped[str]
    entity: Mapped[str]
    entity_id: Mapped[int]
    message: Mapped[str]
    created_at: Mapped[datetime] = mapped_column(default=datetime.now)


engine = create_engine(
    "postgresql+psycopg2://postgres:postgres@localhost:5434/postgres", echo=False
)


@event.listens_for(engine, "connect")
def receive_connect(dbapi_connection, connection_record):
    """Включаем вывод NOTICE сообщений из PostgreSQL в консоль Python."""

    def notice_handler(notice):
        # notice может быть объектом или строкой в зависимости от версии psycopg2
        if hasattr(notice, "message"):
            print(f"DB NOTICE: {notice.message.strip()}")
        else:
            print(f"DB NOTICE: {str(notice).strip()}")

    if hasattr(dbapi_connection, "set_notice_receiver"):
        dbapi_connection.set_notice_receiver(notice_handler)


Session = sessionmaker(bind=engine)
Base.metadata.create_all(engine)


# ===================== TRIGGERS (3 DML & 1 DDL) =====================

# В PostgreSQL триггеры обычно состоят из функции и самого триггера.

# 1. DML Trigger: Обновление общей суммы заказа
trigger_dml_1_func = """
CREATE OR REPLACE FUNCTION update_order_total()
RETURNS TRIGGER AS $$
DECLARE
    target_order_id INT;
BEGIN
    IF (TG_OP = 'DELETE') THEN
        target_order_id = OLD.order_id;
    ELSE
        target_order_id = NEW.order_id;
    END IF;

    UPDATE orders
    SET total_amount = (
        SELECT COALESCE(SUM(quantity * price_at_sale), 0)
        FROM order_compositions
        WHERE order_id = target_order_id
    )
    WHERE id = target_order_id;

    IF (TG_OP = 'DELETE') THEN
        RETURN OLD;
    END IF;
    RETURN NEW;
END;
$$ LANGUAGE plpgsql;
"""

trigger_dml_1 = """
DROP TRIGGER IF EXISTS trg_update_order_total ON order_compositions;
CREATE TRIGGER trg_update_order_total
AFTER INSERT OR UPDATE OR DELETE ON order_compositions
FOR EACH ROW EXECUTE FUNCTION update_order_total();
"""

# 2. DML Trigger: Простое уведомление об изменении клиента
trigger_dml_2_func = """
CREATE OR REPLACE FUNCTION notify_customer_update()
RETURNS TRIGGER AS $$
BEGIN
    INSERT INTO trigger_logs (
        trigger_name,
        action,
        entity,
        entity_id,
        message,
        created_at
    )
    VALUES (
        'trg_notify_customer_update',
        'UPDATE',
        'customers',
        NEW.id,
        'Обновлены данные клиента: ' || NEW.name,
        NOW()
    );

    RETURN NEW;
END;
$$ LANGUAGE plpgsql;

"""

trigger_dml_2 = """
DROP TRIGGER IF EXISTS trg_notify_customer_update ON customers;
CREATE TRIGGER trg_notify_customer_update
AFTER UPDATE ON customers
FOR EACH ROW EXECUTE FUNCTION notify_customer_update();
"""

# 3. DML Trigger: Логирование удаления блюда из меню
trigger_dml_3_func = """
CREATE OR REPLACE FUNCTION notify_menu_item_deletion()
RETURNS TRIGGER AS $$
BEGIN
    INSERT INTO trigger_logs (
        trigger_name,
        action,
        entity,
        entity_id,
        message,
        created_at
    )
    VALUES (
        'trg_notify_menu_item_deletion',
        'DELETE',
        'menu_items',
        OLD.id,
        'Удалено блюдо из меню: ' || OLD.name,
        NOW()
    );

    RETURN OLD;
END;
$$ LANGUAGE plpgsql;

"""

trigger_dml_3 = """
DROP TRIGGER IF EXISTS trg_notify_menu_item_deletion ON menu_items;
CREATE TRIGGER trg_notify_menu_item_deletion
AFTER DELETE ON menu_items
FOR EACH ROW EXECUTE FUNCTION notify_menu_item_deletion();
"""

# 4. DDL Trigger: Простое уведомление о DDL
trigger_ddl_func = """
CREATE OR REPLACE FUNCTION notify_ddl_action()
RETURNS event_trigger AS $$
BEGIN
    RAISE NOTICE 'Выполнена DDL операция: %', TG_TAG;
END;
$$ LANGUAGE plpgsql;
"""

# Event trigger нельзя создать внутри транзакции в некоторых случаях,
# и он требует прав суперпользователя.
# Также он срабатывает на всю базу.
trigger_ddl = """
DROP EVENT TRIGGER IF EXISTS trg_notify_ddl;
CREATE EVENT TRIGGER trg_notify_ddl ON ddl_command_end
EXECUTE FUNCTION notify_ddl_action();
"""

with engine.connect() as conn:
    print("Creating triggers...")
    # Обычные DML/Table DDL можно в транзакции
    conn.execute(text(trigger_dml_1_func))
    conn.execute(text(trigger_dml_1))
    conn.execute(text(trigger_dml_2_func))
    conn.execute(text(trigger_dml_2))
    conn.execute(text(trigger_dml_3_func))
    conn.execute(text(trigger_dml_3))
    conn.execute(text(trigger_ddl_func))
    # Event triggers в PG не могут быть созданы в блоке с другими командами иногда,
    # или требуют отдельного коммита.
    conn.commit()
    print("Triggers created.")

# Event trigger часто требует отдельного выполнения, так как он глобальный.
try:
    with engine.connect() as conn:
        conn.execute(text(trigger_ddl))
        conn.commit()
except Exception as e:
    print(f"Could not create event trigger (might need superuser): {e}")


# ===================== HELPERS =====================


def refresh_order_compositions():
    """
    Принудительно перечитывает OrderComposition из БД
    (учитывает изменения после триггеров)
    """
    s = Session()
    s.expire_all()  # сброс ORM-кэша
    items = s.query(OrderComposition).all()
    rows = [
        (
            c.id,
            c.order_id,
            c.menu_item.name,
            c.quantity,
            c.price_at_sale,
            c.total_price(),
        )
        for c in items
    ]
    s.close()
    reload_tree(compositions_tree, rows)

    # заодно обновим список заказов (total_amount меняется триггером)
    load_orders()


def validate_russian_phone(phone: str) -> bool:
    """
    Валидация российского номера телефона.
    Допустимые форматы:
    +79991234567
    89991234567
    79991234567
    9991234567
    Также допускаются скобки, тире и пробелы, которые будут удалены перед проверкой.
    """
    # Удаляем все кроме цифр и начального плюса
    cleaned = re.sub(r"[^\d+]", "", phone)

    # Регулярное выражение для российского мобильного/городского:
    # Может начинаться с +7, 7, 8 или без префикса (тогда 10 цифр)
    pattern = r"^(\+7|7|8)?[\s\-]?\(?[489][0-9]{2}\)?[\s\-]?[0-9]{3}[\s\-]?[0-9]{2}[\s\-]?[0-9]{2}$"

    # Но проще проверить после очистки:
    digits = re.sub(r"\D", "", cleaned)

    if len(digits) == 10:
        return digits.startswith(("4", "8", "9"))  # упрощенно
    elif len(digits) == 11:
        return digits.startswith(("7", "8"))

    return False


def create_table(parent, columns, headers):
    tree = ttk.Treeview(parent, columns=columns, show="headings")
    for c, h in zip(columns, headers):
        tree.heading(c, text=h)
        tree.column(c, width=150)
    tree.pack(fill="both", expand=True)
    return tree


def reload_tree(tree, rows):
    tree.delete(*tree.get_children())
    for r in rows:
        tree.insert("", "end", values=[str(v) if v is not None else "" for v in r])


def delete_selected(tree, model, reload_func):
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Внимание", "Выберите запись")
        return

    if not messagebox.askyesno("Подтверждение", "Удалить выбранную запись?"):
        return

    item = tree.item(selected[0])
    record_id = int(item["values"][0])

    s = Session()
    obj = s.get(model, record_id)
    if obj:
        s.delete(obj)
        s.commit()
    s.close()

    reload_func()


def get_selected_id(tree):
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Внимание", "Выберите запись")
        return None
    item = tree.item(selected[0])
    return int(item["values"][0])


# ===================== LOADERS =====================


def load_customers():
    s = Session()
    rows = s.query(
        Customer.id,
        Customer.name,
        Customer.phone,
        Customer.email,
        Customer.loyalty_level,
        Customer.discount_percent,
    ).all()
    s.close()
    reload_tree(customers_tree, rows)


def load_logs():
    s = Session()
    rows = (
        s.query(
            TriggerLog.id,
            TriggerLog.created_at,
            TriggerLog.trigger_name,
            TriggerLog.action,
            TriggerLog.entity,
            TriggerLog.entity_id,
            TriggerLog.message,
        )
        .order_by(TriggerLog.created_at.desc())
        .all()
    )
    s.close()
    reload_tree(logs_tree, rows)


def load_suppliers():
    s = Session()
    rows = s.query(
        Supplier.id, Supplier.name, Supplier.phone, Supplier.email, Supplier.address
    ).all()
    s.close()
    reload_tree(suppliers_tree, rows)


def load_ingredients():
    s = Session()
    rows = (
        s.query(
            Ingredient.id,
            Ingredient.name,
            Ingredient.unit,
            Ingredient.stock_quantity,
            Ingredient.purchase_price,
            Supplier.name,
        )
        .join(Supplier)
        .all()
    )

    s.close()
    reload_tree(ingredients_tree, rows)


def load_orders():
    s = Session()

    orders = s.query(Order).all()
    rows = []

    for o in orders:
        rows.append(
            (
                o.id,
                o.customer.name if o.customer else "<В зале>",
                o.employee.fio,
                o.order_date.strftime("%Y-%m-%d %H:%M"),
                len(o.compositions),
                float(o.total_amount),
                o.payment_method,
                o.status,
            )
        )

    s.close()
    reload_tree(orders_tree, rows)


def load_employees():
    s = Session()
    rows = s.query(
        Employee.id, Employee.fio, Employee.role, Employee.phone, Employee.salary
    ).all()
    s.close()
    reload_tree(employees_tree, rows)


def load_menu():
    s = Session()
    rows = s.query(
        MenuItem.id,
        MenuItem.name,
        MenuItem.type,
        MenuItem.selling_price,
        MenuItem.volume_or_weight,
    ).all()
    s.close()
    reload_tree(menu_tree, rows)


def load_recipes():
    s = Session()
    items = s.query(Recipe).all()
    rows = [
        (r.id, r.menu_item.name, r.ingredient.name, r.quantity_required, r.unit)
        for r in items
    ]
    s.close()
    reload_tree(recipes_tree, rows)


def load_order_compositions():
    s = Session()
    items = s.query(OrderComposition).all()
    rows = [
        (
            c.id,
            c.order_id,
            c.menu_item.name,
            c.quantity,
            c.price_at_sale,
            c.total_price(),
        )
        for c in items
    ]
    s.close()
    reload_tree(compositions_tree, rows)


# ===================== CREATE FORMS =====================


def report_all_orders():
    s = Session()

    rows = (
        s.query(
            Order.id,
            Customer.name,
            MenuItem.name,
            OrderComposition.quantity,
            OrderComposition.price_at_sale,
            (OrderComposition.quantity * OrderComposition.price_at_sale).label("total"),
        )
        .outerjoin(Customer, Order.customer_id == Customer.id)
        .join(OrderComposition, OrderComposition.order_id == Order.id)
        .join(MenuItem, OrderComposition.menu_item_id == MenuItem.id)
        .all()
    )

    s.close()
    return rows


def create_order_composition():
    win = tk.Toplevel(root)
    win.title("Добавить позицию в заказ")

    tk.Label(win, text="Заказ ID").grid(row=0, column=0)
    tk.Label(win, text="Позиция").grid(row=1, column=0)
    tk.Label(win, text="Количество").grid(row=2, column=0)
    tk.Label(win, text="Цена").grid(row=3, column=0)

    order_var = tk.StringVar()
    item_var = tk.StringVar()

    order_box = ttk.Combobox(win, textvariable=order_var, state="readonly")
    item_box = ttk.Combobox(win, textvariable=item_var, state="readonly")

    e_qty = tk.Entry(win)
    e_price = tk.Entry(win)

    order_box.grid(row=0, column=1)
    item_box.grid(row=1, column=1)
    e_qty.grid(row=2, column=1)
    e_price.grid(row=3, column=1)

    s = Session()
    orders = s.query(Order).all()
    items = s.query(MenuItem).all()

    order_map = {str(o.id): o.id for o in orders}
    item_map = {m.name: m.id for m in items}

    order_box["values"] = list(order_map.keys())
    item_box["values"] = list(item_map.keys())
    s.close()

    def save():
        s = Session()
        s.add(
            OrderComposition(
                order_id=order_map[order_var.get()],
                menu_item_id=item_map[item_var.get()],
                quantity=int(e_qty.get()),
                price_at_sale=float(e_price.get()),
            )
        )
        s.commit()
        s.close()
        load_order_compositions()
        load_orders()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def edit_order_composition(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return

    s = Session()
    c = s.get(OrderComposition, record_id)

    win = tk.Toplevel(root)
    win.title("Редактировать позицию")

    tk.Label(win, text="Количество").grid(row=0, column=0)
    tk.Label(win, text="Цена").grid(row=1, column=0)

    e_qty = tk.Entry(win)
    e_qty.insert(0, c.quantity)
    e_qty.grid(row=0, column=1)

    e_price = tk.Entry(win)
    e_price.insert(0, c.price_at_sale)
    e_price.grid(row=1, column=1)

    def save():
        c.quantity = int(e_qty.get())
        c.price_at_sale = float(e_price.get())
        s.commit()
        s.close()
        load_order_compositions()
        load_orders()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_order():
    win = tk.Toplevel(root)
    win.title("Новый заказ")
    win.geometry("500x600")

    tk.Label(win, text="Сотрудник (принял заказ)").pack()
    employee_var = tk.StringVar()
    employee_box = ttk.Combobox(win, textvariable=employee_var, state="readonly")
    employee_box.pack(fill="x")

    tk.Label(win, text="Клиент (необязательно)").pack()
    customer_var = tk.StringVar()
    customer_box = ttk.Combobox(win, textvariable=customer_var, state="readonly")
    customer_box.pack(fill="x")

    tk.Label(win, text="Тип заказа (В зале/С собой)").pack()
    type_entry = tk.Entry(win)
    type_entry.insert(0, "В зале")
    type_entry.pack(fill="x")

    tk.Label(win, text="Метод оплаты").pack()
    payment_entry = tk.Entry(win)
    payment_entry.insert(0, "Карта")
    payment_entry.pack(fill="x")

    tk.Label(win, text="Позиции").pack()

    menu_list = tk.Listbox(win, height=10)
    menu_list.pack(fill="both", expand=True)

    tk.Label(win, text="Количество").pack()
    qty = tk.Entry(win)
    qty.insert(0, "1")
    qty.pack()

    cart = []

    cart_box = tk.Listbox(win, height=6)
    cart_box.pack(fill="both", expand=True)

    s = Session()

    employees = s.query(Employee).all()
    employee_map = {e.fio: e.id for e in employees}
    employee_box["values"] = list(employee_map.keys())

    customers = s.query(Customer).all()
    customer_map = {c.name: c.id for c in customers}
    customer_box["values"] = ["<Нет клиента>"] + list(customer_map.keys())
    customer_box.set("<Нет клиента>")

    items = s.query(MenuItem).all()
    item_map = {f"{i.name} ({i.selling_price})": (i.id, i.selling_price) for i in items}
    menu_list.insert("end", *item_map.keys())

    s.close()

    def add_to_cart():
        sel = menu_list.curselection()
        if not sel:
            return

        key = menu_list.get(sel)
        item_id, price = item_map[key]
        try:
            quantity = int(qty.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Количество должно быть числом")
            return

        cart.append((item_id, quantity, price))
        cart_box.insert("end", f"{key} x{quantity}")

    def save():
        if not employee_var.get():
            messagebox.showerror("Ошибка", "Выберите сотрудника")
            return

        if not cart:
            messagebox.showerror("Ошибка", "Корзина пуста")
            return

        s = Session()

        c_id = None
        if customer_var.get() and customer_var.get() != "<Нет клиента>":
            c_id = customer_map[customer_var.get()]

        try:
            order = Order(
                customer_id=c_id,
                employee_id=employee_map[employee_var.get()],
                order_type=type_entry.get(),
                payment_method=payment_entry.get(),
                status="Новый",
            )
            s.add(order)
            s.flush()

            for item_id, quantity, price in cart:
                s.add(
                    OrderComposition(
                        order_id=order.id,
                        menu_item_id=item_id,
                        quantity=quantity,
                        price_at_sale=price,
                    )
                )

            s.commit()
            load_orders()
            win.destroy()
        except Exception as e:
            s.rollback()
            messagebox.showerror("Ошибка", f"Не удалось сохранить заказ: {e}")
        finally:
            s.close()

    tk.Button(win, text="➕ Добавить в чек", command=add_to_cart).pack()
    tk.Button(win, text="💾 Сохранить заказ", command=save).pack()


def create_supplier():
    win = tk.Toplevel(root)
    win.title("Новый поставщик")

    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Телефон").grid(row=1, column=0)
    tk.Label(win, text="Email").grid(row=2, column=0)
    tk.Label(win, text="Адрес").grid(row=3, column=0)

    e_name = tk.Entry(win)
    e_phone = tk.Entry(win)
    e_email = tk.Entry(win)
    e_address = tk.Entry(win)

    e_name.grid(row=0, column=1)
    e_phone.grid(row=1, column=1)
    e_email.grid(row=2, column=1)
    e_address.grid(row=3, column=1)

    def save():
        if not e_name.get():
            messagebox.showerror("Ошибка", "Заполните название")
            return

        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return

        s = Session()
        s.add(
            Supplier(
                name=e_name.get(),
                phone=phone,
                email=e_email.get(),
                address=e_address.get(),
            )
        )
        s.commit()
        s.close()
        load_suppliers()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_ingredient():
    win = tk.Toplevel(root)
    win.title("Новый ингредиент")

    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Ед. изм.").grid(row=1, column=0)
    tk.Label(win, text="Кол-во").grid(row=2, column=0)
    tk.Label(win, text="Мин. уровень").grid(row=3, column=0)
    tk.Label(win, text="Цена закупки").grid(row=4, column=0)
    tk.Label(win, text="Поставщик").grid(row=5, column=0)

    e_name = tk.Entry(win)
    e_unit = tk.Entry(win)
    e_qty = tk.Entry(win)
    e_min = tk.Entry(win)
    e_price = tk.Entry(win)

    e_name.grid(row=0, column=1)
    e_unit.grid(row=1, column=1)
    e_qty.grid(row=2, column=1)
    e_min.grid(row=3, column=1)
    e_price.grid(row=4, column=1)

    sup_var = tk.StringVar()
    sup_box = ttk.Combobox(win, textvariable=sup_var, state="readonly")
    sup_box.grid(row=5, column=1)

    s = Session()
    suppliers = s.query(Supplier).all()
    sup_map = {sup.name: sup.id for sup in suppliers}
    sup_box["values"] = list(sup_map.keys())
    s.close()

    def save():
        if not e_name.get() or not sup_var.get():
            messagebox.showerror("Ошибка", "Заполните поля")
            return

        try:
            qty = float(e_qty.get() or 0)
            min_lvl = float(e_min.get() or 0)
            price = float(e_price.get() or 0)
        except ValueError:
            messagebox.showerror(
                "Ошибка",
                "Некорректный тип данных. Поля 'Кол-во', 'Мин. уровень' и 'Цена' должны быть числовыми.",
            )
            return

        s = Session()
        s.add(
            Ingredient(
                name=e_name.get(),
                unit=e_unit.get(),
                stock_quantity=qty,
                min_stock_level=min_lvl,
                purchase_price=price,
                supplier_id=sup_map[sup_var.get()],
            )
        )
        s.commit()
        s.close()
        load_ingredients()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_employee():
    win = tk.Toplevel(root)
    win.title("Новый сотрудник")

    tk.Label(win, text="ФИО").grid(row=0, column=0)
    tk.Label(win, text="Роль").grid(row=1, column=0)
    tk.Label(win, text="Телефон").grid(row=2, column=0)
    tk.Label(win, text="Зарплата").grid(row=3, column=0)

    e_fio = tk.Entry(win)
    e_role = tk.Entry(win)
    e_phone = tk.Entry(win)
    e_salary = tk.Entry(win)

    e_fio.grid(row=0, column=1)
    e_role.grid(row=1, column=1)
    e_phone.grid(row=2, column=1)
    e_salary.grid(row=3, column=1)

    def save():
        try:
            salary = float(e_salary.get() or 0)
        except ValueError:
            messagebox.showerror(
                "Ошибка",
                "Некорректный тип данных. Поле 'Зарплата' должно быть числовым.",
            )
            return

        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return

        s = Session()
        s.add(Employee(fio=e_fio.get(), role=e_role.get(), phone=phone, salary=salary))
        s.commit()
        s.close()
        load_employees()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_menu_item():
    win = tk.Toplevel(root)
    win.title("Новое блюдо")

    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Тип").grid(row=1, column=0)
    tk.Label(win, text="Цена продажи").grid(row=2, column=0)
    tk.Label(win, text="Объем/Вес").grid(row=3, column=0)

    e_name = tk.Entry(win)
    e_type = tk.Entry(win)
    e_price = tk.Entry(win)
    e_vol = tk.Entry(win)

    e_name.grid(row=0, column=1)
    e_type.grid(row=1, column=1)
    e_price.grid(row=2, column=1)
    e_vol.grid(row=3, column=1)

    def save():
        try:
            price = float(e_price.get() or 0)
        except ValueError:
            messagebox.showerror(
                "Ошибка",
                "Некорректный тип данных. Поле 'Цена продажи' должно быть числовым.",
            )
            return

        s = Session()
        s.add(
            MenuItem(
                name=e_name.get(),
                type=e_type.get(),
                selling_price=price,
                volume_or_weight=e_vol.get(),
            )
        )
        s.commit()
        s.close()
        load_menu()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_customer():
    win = tk.Toplevel(root)
    win.title("Новый клиент")

    tk.Label(win, text="Имя").grid(row=0, column=0)
    tk.Label(win, text="Телефон").grid(row=1, column=0)
    tk.Label(win, text="Email").grid(row=2, column=0)

    e_name = tk.Entry(win)
    e_phone = tk.Entry(win)
    e_email = tk.Entry(win)

    e_name.grid(row=0, column=1)
    e_phone.grid(row=1, column=1)
    e_email.grid(row=2, column=1)

    def save():
        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return

        s = Session()
        s.add(Customer(name=e_name.get(), phone=phone, email=e_email.get()))
        s.commit()
        s.close()
        load_customers()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def edit_customer(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    obj = s.get(Customer, record_id)
    if not obj:
        s.close()
        return
    win = tk.Toplevel(root)
    win.title("Редактировать клиента")
    tk.Label(win, text="Имя").grid(row=0, column=0)
    tk.Label(win, text="Телефон").grid(row=1, column=0)
    tk.Label(win, text="Email").grid(row=2, column=0)

    tk.Label(win, text="Уровень").grid(row=3, column=0)
    tk.Label(win, text="Скидка (%)").grid(row=4, column=0)

    e_level = tk.Entry(win)
    e_level.insert(0, obj.loyalty_level)
    e_level.grid(row=3, column=1)

    e_discount = tk.Entry(win)
    e_discount.insert(0, str(obj.discount_percent))
    e_discount.grid(row=4, column=1)

    e_name = tk.Entry(win)
    e_name.insert(0, obj.name)
    e_name.grid(row=0, column=1)
    e_phone = tk.Entry(win)
    e_phone.insert(0, obj.phone)
    e_phone.grid(row=1, column=1)
    e_email = tk.Entry(win)
    e_email.insert(0, obj.email)
    e_email.grid(row=2, column=1)

    def save():
        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return
        s2 = Session()
        o = s2.get(Customer, record_id)
        o.name = e_name.get()
        o.phone = phone
        o.email = e_email.get()

        try:
            discount = float(e_discount.get() or 0)
        except ValueError:
            messagebox.showerror("Ошибка", "Скидка должна быть числом")
            return

        o.loyalty_level = e_level.get()
        o.discount_percent = discount

        s2.commit()
        s2.close()
        load_customers()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)
    s.close()


def edit_employee(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    obj = s.get(Employee, record_id)
    if not obj:
        s.close()
        return
    win = tk.Toplevel(root)
    win.title("Редактировать сотрудника")
    tk.Label(win, text="ФИО").grid(row=0, column=0)
    tk.Label(win, text="Роль").grid(row=1, column=0)
    tk.Label(win, text="Телефон").grid(row=2, column=0)
    tk.Label(win, text="Зарплата").grid(row=3, column=0)
    e_fio = tk.Entry(win)
    e_fio.insert(0, obj.fio)
    e_fio.grid(row=0, column=1)
    e_role = tk.Entry(win)
    e_role.insert(0, obj.role)
    e_role.grid(row=1, column=1)
    e_phone = tk.Entry(win)
    e_phone.insert(0, obj.phone)
    e_phone.grid(row=2, column=1)
    e_salary = tk.Entry(win)
    e_salary.insert(0, str(obj.salary))
    e_salary.grid(row=3, column=1)

    def save():
        try:
            salary = float(e_salary.get() or 0)
        except ValueError:
            messagebox.showerror(
                "Ошибка",
                "Некорректный тип данных. Поле 'Зарплата' должно быть числовым.",
            )
            return
        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return
        s2 = Session()
        o = s2.get(Employee, record_id)
        o.fio = e_fio.get()
        o.role = e_role.get()
        o.phone = phone
        o.salary = salary
        s2.commit()
        s2.close()
        load_employees()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)
    s.close()


def edit_supplier(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    obj = s.get(Supplier, record_id)
    if not obj:
        s.close()
        return
    win = tk.Toplevel(root)
    win.title("Редактировать поставщика")
    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Телефон").grid(row=1, column=0)
    tk.Label(win, text="Email").grid(row=2, column=0)
    tk.Label(win, text="Адрес").grid(row=3, column=0)
    e_name = tk.Entry(win)
    e_name.insert(0, obj.name)
    e_name.grid(row=0, column=1)
    e_phone = tk.Entry(win)
    e_phone.insert(0, obj.phone)
    e_phone.grid(row=1, column=1)
    e_email = tk.Entry(win)
    e_email.insert(0, obj.email)
    e_email.grid(row=2, column=1)
    e_address = tk.Entry(win)
    e_address.insert(0, obj.address)
    e_address.grid(row=3, column=1)

    def save():
        if not e_name.get():
            messagebox.showerror("Ошибка", "Заполните название")
            return
        phone = e_phone.get()
        if phone and not validate_russian_phone(phone):
            messagebox.showerror("Ошибка", "Некорректный российский номер телефона")
            return
        s2 = Session()
        o = s2.get(Supplier, record_id)
        o.name = e_name.get()
        o.phone = phone
        o.email = e_email.get()
        o.address = e_address.get()
        s2.commit()
        s2.close()
        load_suppliers()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)
    s.close()


def edit_ingredient(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    obj = s.get(Ingredient, record_id)
    if not obj:
        s.close()
        return
    win = tk.Toplevel(root)
    win.title("Редактировать ингредиент")
    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Ед. изм.").grid(row=1, column=0)
    tk.Label(win, text="Кол-во").grid(row=2, column=0)
    tk.Label(win, text="Мин. уровень").grid(row=3, column=0)
    tk.Label(win, text="Цена закупки").grid(row=4, column=0)
    tk.Label(win, text="Поставщик").grid(row=5, column=0)
    e_name = tk.Entry(win)
    e_name.insert(0, obj.name)
    e_name.grid(row=0, column=1)
    e_unit = tk.Entry(win)
    e_unit.insert(0, obj.unit)
    e_unit.grid(row=1, column=1)
    e_qty = tk.Entry(win)
    e_qty.insert(0, str(obj.stock_quantity))
    e_qty.grid(row=2, column=1)
    e_min = tk.Entry(win)
    e_min.insert(0, str(obj.min_stock_level))
    e_min.grid(row=3, column=1)
    e_price = tk.Entry(win)
    e_price.insert(0, str(obj.purchase_price))
    e_price.grid(row=4, column=1)
    sup_var = tk.StringVar()
    sup_box = ttk.Combobox(win, textvariable=sup_var, state="readonly")
    sup_box.grid(row=5, column=1)
    suppliers = s.query(Supplier).all()
    sup_map = {sup.name: sup.id for sup in suppliers}
    sup_box["values"] = list(sup_map.keys())
    for name, sid in sup_map.items():
        if sid == obj.supplier_id:
            sup_box.set(name)
            break

    def save():
        if not e_name.get() or not sup_var.get():
            messagebox.showerror("Ошибка", "Заполните поля")
            return
        try:
            qty = float(e_qty.get() or 0)
            min_lvl = float(e_min.get() or 0)
            price = float(e_price.get() or 0)
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректный тип данных.")
            return
        s2 = Session()
        o = s2.get(Ingredient, record_id)
        o.name = e_name.get()
        o.unit = e_unit.get()
        o.stock_quantity = qty
        o.min_stock_level = min_lvl
        o.purchase_price = price
        o.supplier_id = sup_map[sup_var.get()]
        s2.commit()
        s2.close()
        load_ingredients()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)
    s.close()


def edit_menu_item(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    obj = s.get(MenuItem, record_id)
    if not obj:
        s.close()
        return
    win = tk.Toplevel(root)
    win.title("Редактировать блюдо")
    tk.Label(win, text="Название").grid(row=0, column=0)
    tk.Label(win, text="Тип").grid(row=1, column=0)
    tk.Label(win, text="Цена продажи").grid(row=2, column=0)
    tk.Label(win, text="Объем/Вес").grid(row=3, column=0)
    e_name = tk.Entry(win)
    e_name.insert(0, obj.name)
    e_name.grid(row=0, column=1)
    e_type = tk.Entry(win)
    e_type.insert(0, obj.type)
    e_type.grid(row=1, column=1)
    e_price = tk.Entry(win)
    e_price.insert(0, str(obj.selling_price))
    e_price.grid(row=2, column=1)
    e_vol = tk.Entry(win)
    e_vol.insert(0, obj.volume_or_weight)
    e_vol.grid(row=3, column=1)

    def save():
        try:
            price = float(e_price.get() or 0)
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректный тип данных.")
            return
        s2 = Session()
        o = s2.get(MenuItem, record_id)
        o.name = e_name.get()
        o.type = e_type.get()
        o.selling_price = price
        o.volume_or_weight = e_vol.get()
        s2.commit()
        s2.close()
        load_menu()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)
    s.close()


def edit_recipe(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return

    s = Session()
    r = s.get(Recipe, record_id)

    win = tk.Toplevel(root)
    win.title("Редактировать рецепт")

    tk.Label(win, text="Количество").grid(row=0, column=0)
    tk.Label(win, text="Ед. изм.").grid(row=1, column=0)

    e_qty = tk.Entry(win)
    e_qty.insert(0, r.quantity_required)
    e_qty.grid(row=0, column=1)

    e_unit = tk.Entry(win)
    e_unit.insert(0, r.unit)
    e_unit.grid(row=1, column=1)

    def save():
        try:
            r.quantity_required = float(e_qty.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Количество должно быть числом")
            return

        r.unit = e_unit.get()
        s.commit()
        s.close()
        load_recipes()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def create_recipe():
    win = tk.Toplevel(root)
    win.title("Новый рецепт")

    tk.Label(win, text="Блюдо").grid(row=0, column=0)
    tk.Label(win, text="Ингредиент").grid(row=1, column=0)
    tk.Label(win, text="Количество").grid(row=2, column=0)
    tk.Label(win, text="Ед. изм.").grid(row=3, column=0)

    menu_var = tk.StringVar()
    ing_var = tk.StringVar()

    menu_box = ttk.Combobox(win, textvariable=menu_var, state="readonly")
    ing_box = ttk.Combobox(win, textvariable=ing_var, state="readonly")

    e_qty = tk.Entry(win)
    e_unit = tk.Entry(win)

    menu_box.grid(row=0, column=1)
    ing_box.grid(row=1, column=1)
    e_qty.grid(row=2, column=1)
    e_unit.grid(row=3, column=1)

    s = Session()
    menu = s.query(MenuItem).all()
    ingredients = s.query(Ingredient).all()

    menu_map = {m.name: m.id for m in menu}
    ing_map = {i.name: i.id for i in ingredients}

    menu_box["values"] = list(menu_map.keys())
    ing_box["values"] = list(ing_map.keys())
    s.close()

    def save():
        try:
            qty = float(e_qty.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Количество должно быть числом")
            return

        s = Session()
        s.add(
            Recipe(
                menu_item_id=menu_map[menu_var.get()],
                ingredient_id=ing_map[ing_var.get()],
                quantity_required=qty,
                unit=e_unit.get(),
            )
        )
        s.commit()
        s.close()
        load_recipes()
        win.destroy()

    tk.Button(win, text="Сохранить", command=save).grid(columnspan=2)


def edit_order(tree):
    record_id = get_selected_id(tree)
    if record_id is None:
        return
    s = Session()
    order = s.query(Order).filter(Order.id == record_id).first()
    if not order:
        s.close()
        return

    win = tk.Toplevel(root)
    win.title(f"Редактировать заказ №{record_id}")
    win.geometry("500x700")

    tk.Label(win, text="Сотрудник").pack()
    employee_var = tk.StringVar()
    employee_box = ttk.Combobox(win, textvariable=employee_var, state="readonly")
    employee_box.pack(fill="x")

    tk.Label(win, text="Клиент").pack()
    customer_var = tk.StringVar()
    customer_box = ttk.Combobox(win, textvariable=customer_var, state="readonly")
    customer_box.pack(fill="x")

    tk.Label(win, text="Тип заказа").pack()
    type_entry = tk.Entry(win)
    type_entry.insert(0, order.order_type)
    type_entry.pack(fill="x")

    tk.Label(win, text="Метод оплаты").pack()
    payment_entry = tk.Entry(win)
    payment_entry.insert(0, order.payment_method)
    payment_entry.pack(fill="x")

    tk.Label(win, text="Статус").pack()
    status_entry = tk.Entry(win)
    status_entry.insert(0, order.status)
    status_entry.pack(fill="x")

    tk.Label(win, text="Добавить позиции").pack()
    menu_list = tk.Listbox(win, height=8)
    menu_list.pack(fill="both", expand=True)

    tk.Label(win, text="Количество").pack()
    qty = tk.Entry(win)
    qty.insert(0, "1")
    qty.pack()

    cart = []
    for comp in order.compositions:
        cart.append(
            {
                "id": comp.menu_item_id,
                "name": comp.menu_item.name,
                "quantity": comp.quantity,
                "price": comp.price_at_sale,
            }
        )

    cart_box = tk.Listbox(win, height=6)
    cart_box.pack(fill="both", expand=True)

    def refresh_cart_box():
        cart_box.delete(0, "end")
        for i in cart:
            cart_box.insert("end", f"{i['name']} ({i['price']}) x{i['quantity']}")

    refresh_cart_box()

    employees = s.query(Employee).all()
    employee_map = {e.fio: e.id for e in employees}
    employee_box["values"] = list(employee_map.keys())
    for name, eid in employee_map.items():
        if eid == order.employee_id:
            employee_box.set(name)
            break

    customers = s.query(Customer).all()
    customer_map = {c.name: c.id for c in customers}
    customer_box["values"] = ["<Нет клиента>"] + list(customer_map.keys())
    if order.customer:
        customer_box.set(order.customer.name)
    else:
        customer_box.set("<Нет клиента>")

    items = s.query(MenuItem).all()
    item_map = {
        f"{i.name} ({i.selling_price})": (i.id, i.selling_price, i.name) for i in items
    }
    menu_list.insert("end", *item_map.keys())

    def add_to_cart():
        sel = menu_list.curselection()
        if not sel:
            return
        key = menu_list.get(sel)
        item_id, price, name = item_map[key]
        try:
            quantity = int(qty.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Количество должно быть числом")
            return

        # Если такой товар уже есть, прибавляем
        for i in cart:
            if i["id"] == item_id:
                i["quantity"] += quantity
                refresh_cart_box()
                return

        cart.append({"id": item_id, "name": name, "quantity": quantity, "price": price})
        refresh_cart_box()

    def remove_from_cart():
        sel = cart_box.curselection()
        if not sel:
            return
        cart.pop(sel[0])
        refresh_cart_box()

    def save():
        if not employee_var.get():
            messagebox.showerror("Ошибка", "Выберите сотрудника")
            return
        if not cart:
            messagebox.showerror("Ошибка", "Корзина пуста")
            return

        s2 = Session()
        try:
            o = s2.query(Order).filter(Order.id == record_id).first()
            o.employee_id = employee_map[employee_var.get()]
            c_id = None
            if customer_var.get() != "<Нет клиента>":
                c_id = customer_map[customer_var.get()]
            o.customer_id = c_id
            o.order_type = type_entry.get()
            o.payment_method = payment_entry.get()
            o.status = status_entry.get()

            # Обновляем состав: проще всего удалить старые и добавить новые
            s2.query(OrderComposition).filter(
                OrderComposition.order_id == record_id
            ).delete()
            for i in cart:
                s2.add(
                    OrderComposition(
                        order_id=record_id,
                        menu_item_id=i["id"],
                        quantity=i["quantity"],
                        price_at_sale=i["price"],
                    )
                )

            s2.commit()
            load_orders()
            win.destroy()
        except Exception as e:
            s2.rollback()
            messagebox.showerror("Ошибка", f"Не удалось сохранить: {e}")
        finally:
            s2.close()

    tk.Button(win, text="➕ Добавить", command=add_to_cart).pack(side="top")
    tk.Button(win, text="❌ Удалить позицию", command=remove_from_cart).pack(side="top")
    tk.Button(
        win, text="💾 Сохранить изменения", command=save, bg="green", fg="white"
    ).pack(pady=10)

    s.close()


def export_report():
    s = Session()
    wb = Workbook()

    def bold(ws):
        for c in ws[1]:
            c.font = Font(bold=True)

    # =====================================================
    # 1. ПРОДАЖИ ПО ДНЯМ
    # =====================================================
    ws = wb.active
    ws.title = "Продажи по дням"
    ws.append(["Дата", "Заказов", "Позиций", "Выручка"])

    rows = s.execute(
        text("""
        SELECT
            DATE(o.order_date) AS day,
            COUNT(DISTINCT o.id) AS orders,
            SUM(oc.quantity) AS items,
            SUM(oc.quantity * oc.price_at_sale) AS revenue
        FROM orders o
        JOIN order_compositions oc ON oc.order_id = o.id
        GROUP BY day
        ORDER BY day
    """)
    ).all()

    for r in rows:
        ws.append(tuple(r))

    bold(ws)

    # =====================================================
    # 2. ТОП БЛЮД
    # =====================================================
    ws = wb.create_sheet("Топ блюд")
    ws.append(["Блюдо", "Тип", "Продано", "Выручка"])

    rows = s.execute(
        text("""
        SELECT
            mi.name,
            mi.type,
            SUM(oc.quantity) AS total_qty,
            SUM(oc.quantity * oc.price_at_sale) AS total_revenue
        FROM order_compositions oc
        JOIN menu_items mi ON mi.id = oc.menu_item_id
        GROUP BY mi.id, mi.name, mi.type
        ORDER BY total_revenue DESC
    """)
    ).all()

    for r in rows:
        ws.append(tuple(r))

    bold(ws)

    # =====================================================
    # 3. КЛИЕНТЫ (LTV)
    # =====================================================
    ws = wb.create_sheet("Клиенты LTV")
    ws.append(["Клиент", "Email", "Заказов", "Сумма", "Средний чек"])

    rows = s.execute(
        text("""
        SELECT
            c.name,
            c.email,
            COUNT(o.id),
            COALESCE(SUM(oc.quantity * oc.price_at_sale), 0),
            COALESCE(
                SUM(oc.quantity * oc.price_at_sale) / NULLIF(COUNT(o.id), 0),
                0
            )
        FROM customers c
        LEFT JOIN orders o ON o.customer_id = c.id
        LEFT JOIN order_compositions oc ON oc.order_id = o.id
        GROUP BY c.name, c.email
    """)
    ).all()

    for r in rows:
        ws.append(tuple(r))

    bold(ws)

    # =====================================================
    # 4. ЭФФЕКТИВНОСТЬ СОТРУДНИКОВ
    # =====================================================
    ws = wb.create_sheet("Эффективность")
    ws.append(["Сотрудник", "Заказов", "Сумма заказов"])

    rows = s.execute(
        text("""
        SELECT
            e.fio,
            COUNT(o.id),
            SUM(o.total_amount)
        FROM orders o
        JOIN employees e ON e.id = o.employee_id
        GROUP BY e.fio
    """)
    ).all()

    for r in rows:
        ws.append(tuple(r))

    bold(ws)

    # =====================================================
    # 5. ПРИБЫЛЬ И УБЫТКИ
    # =====================================================
    ws = wb.create_sheet("Прибыль")
    ws.append(["Показатель", "Значение"])

    # 5.1 Выручка (уже посчитана выше, но возьмем для чистоты)
    revenue = s.execute(
        text("""
        SELECT COALESCE(SUM(total_amount), 0) FROM orders
    """)
    ).scalar()

    # 5.2 Себестоимость (COGS) - на основе рецептов и закупочных цен
    cogs = s.execute(
        text("""
        SELECT COALESCE(SUM(oc.quantity * r.quantity_required * i.purchase_price), 0)
        FROM order_compositions oc
        JOIN recipes r ON oc.menu_item_id = r.menu_item_id
        JOIN ingredients i ON r.ingredient_id = i.id
    """)
    ).scalar()

    # 5.3 Расходы на персонал (сумма всех зарплат)
    salaries = s.execute(
        text("SELECT COALESCE(SUM(salary), 0) FROM employees")
    ).scalar()

    profit = float(revenue) - float(cogs) - float(salaries)

    ws.append(["Общая выручка", revenue])
    ws.append(["Себестоимость товаров (продукты)", cogs])
    ws.append(["Расходы на персонал (зарплаты)", salaries])
    ws.append(["Чистая прибыль", profit])

    # Добавим форматирование
    bold(ws)
    for row in ws.iter_rows(min_row=2, max_row=5, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "#,##0.00"

    # =====================================================
    # 6. НАГРУЗКА ПО ЧАСАМ
    # =====================================================
    ws = wb.create_sheet("Нагрузка по часам")
    ws.append(["Час", "Заказов", "Выручка"])

    rows = s.execute(
        text("""
        SELECT
            EXTRACT(HOUR FROM order_date),
            COUNT(*),
            SUM(total_amount)
        FROM orders
        GROUP BY 1
        ORDER BY 1
    """)
    ).all()

    for h, cnt, money in rows:
        ws.append([f"{int(h)}:00", cnt, money])

    bold(ws)

    s.close()

    # =====================================================
    # SAVE
    # =====================================================
    file = asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Сохранить отчет",
    )

    if file:
        wb.save(file)
        messagebox.showinfo("Готово", "Отчёт успешно сохранён!")


# ===================== GUI =====================

root = tk.Tk()
root.title("Cafe Control")
root.geometry("1280x720")

nb = ttk.Notebook(root)
nb.pack(fill="both", expand=True)
tk.Button(
    root,
    text="Выгрузить отчеты в Excel",
    font=("Arial", 12, "bold"),
    command=export_report,
).pack(fill="x", padx=10, pady=5)

# Customers
fc = ttk.Frame(nb)
nb.add(fc, text="Customers")
customers_tree = create_table(
    fc,
    ("id", "name", "phone", "email", "loyalty", "discount"),
    ("ID", "Имя", "Телефон", "Email", "Уровень", "Скидка (%)"),
)
tk.Button(fc, text="Добавить", command=create_customer).pack(side="left")
tk.Button(fc, text="Редактировать", command=lambda: edit_customer(customers_tree)).pack(
    side="left"
)
tk.Button(
    fc,
    text="Удалить",
    command=lambda: delete_selected(customers_tree, Customer, load_customers),
).pack(side="left")
load_customers()

# Employees
fe = ttk.Frame(nb)
nb.add(fe, text="Employees")
employees_tree = create_table(
    fe,
    ("id", "fio", "role", "phone", "salary"),
    ("ID", "ФИО", "Роль", "Телефон", "Зарплата"),
)
tk.Button(fe, text="Добавить", command=create_employee).pack(side="left")
tk.Button(fe, text="Редактировать", command=lambda: edit_employee(employees_tree)).pack(
    side="left"
)
tk.Button(
    fe,
    text="Удалить",
    command=lambda: delete_selected(employees_tree, Employee, load_employees),
).pack(side="left")
load_employees()

# Suppliers
fsup = ttk.Frame(nb)
nb.add(fsup, text="Suppliers")
suppliers_tree = create_table(
    fsup,
    ("id", "name", "phone", "email", "address"),
    ("ID", "Название", "Телефон", "Email", "Адрес"),
)
tk.Button(fsup, text="Добавить", command=create_supplier).pack(side="left")
tk.Button(
    fsup, text="Редактировать", command=lambda: edit_supplier(suppliers_tree)
).pack(side="left")
tk.Button(
    fsup,
    text="Удалить",
    command=lambda: delete_selected(suppliers_tree, Supplier, load_suppliers),
).pack(side="left")
load_suppliers()

# Ingredients
fing = ttk.Frame(nb)
nb.add(fing, text="Ingredients")
ingredients_tree = create_table(
    fing,
    ("id", "name", "unit", "stock", "price", "supplier"),
    ("ID", "Название", "Ед. изм.", "Остаток", "Цена зак.", "Поставщик"),
)
tk.Button(fing, text="Добавить", command=create_ingredient).pack(side="left")
tk.Button(
    fing, text="Редактировать", command=lambda: edit_ingredient(ingredients_tree)
).pack(side="left")
tk.Button(
    fing,
    text="Удалить",
    command=lambda: delete_selected(ingredients_tree, Ingredient, load_ingredients),
).pack(side="left")
load_ingredients()

# Menu
fm = ttk.Frame(nb)
nb.add(fm, text="Menu")
menu_tree = create_table(
    fm,
    ("id", "name", "type", "price", "vol"),
    ("ID", "Название", "Тип", "Цена прод.", "Объем/Вес"),
)
tk.Button(fm, text="Добавить", command=create_menu_item).pack(side="left")
tk.Button(fm, text="Редактировать", command=lambda: edit_menu_item(menu_tree)).pack(
    side="left"
)
tk.Button(
    fm,
    text="Удалить",
    command=lambda: delete_selected(menu_tree, MenuItem, load_menu),
).pack(side="left")
load_menu()

# Orders
fo = ttk.Frame(nb)
nb.add(fo, text="Orders")

orders_tree = create_table(
    fo,
    ("id", "customer", "employee", "date", "items", "total", "payment", "status"),
    ("ID", "Клиент", "Сотрудник", "Дата", "Кол-во", "Сумма", "Оплата", "Статус"),
)
tk.Button(fo, text="Добавить", command=create_order).pack(side="left")
tk.Button(fo, text="Редактировать", command=lambda: edit_order(orders_tree)).pack(
    side="left"
)
tk.Button(
    fo,
    text="Удалить",
    command=lambda: delete_selected(orders_tree, Order, load_orders),
).pack(side="left")
load_orders()

# Recipes
fr = ttk.Frame(nb)
nb.add(fr, text="Recipes")
recipes_tree = create_table(
    fr,
    ("id", "menu_item", "ingredient", "qty", "unit"),
    ("ID", "Блюдо", "Ингредиент", "Кол-во", "Ед. изм."),
)
tk.Button(fr, text="Добавить", command=create_recipe).pack(side="left")
tk.Button(fr, text="Редактировать", command=lambda: edit_recipe(recipes_tree)).pack(
    side="left"
)
tk.Button(
    fr,
    text="Удалить",
    command=lambda: delete_selected(recipes_tree, Recipe, load_recipes),
).pack(side="left")
load_recipes()

# Order Compositions
foc = ttk.Frame(nb)
nb.add(foc, text="Order Details")
compositions_tree = create_table(
    foc,
    ("id", "order_id", "menu_item", "qty", "price", "total"),
    ("ID", "ID Заказа", "Позиция", "Кол-во", "Цена", "Итого"),
)
tk.Button(foc, text="Добавить", command=create_order_composition).pack(side="left")
tk.Button(
    foc,
    text="Редактировать",
    command=lambda: edit_order_composition(compositions_tree),
).pack(side="left")
tk.Button(
    foc,
    text="Удалить",
    command=lambda: delete_selected(
        compositions_tree, OrderComposition, load_order_compositions
    ),
).pack(side="left")
tk.Button(
    foc,
    text="Обновить",
    command=refresh_order_compositions,
).pack(side="left")

# Logs
fl = ttk.Frame(nb)
nb.add(fl, text="Logs")

logs_tree = create_table(
    fl,
    ("id", "time", "trigger", "action", "entity", "entity_id", "message"),
    (
        "ID",
        "Дата",
        "Триггер",
        "Действие",
        "Таблица",
        "ID записи",
        "Сообщение",
    ),
)

tk.Button(fl, text="🔄 Обновить", command=load_logs).pack(side="left")

load_logs()


load_order_compositions()

root.mainloop()
